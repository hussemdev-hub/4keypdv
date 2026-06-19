from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service
import tkinter as tk
from tkinter import messagebox, scrolledtext
import openpyxl
from datetime import datetime, timedelta
import os
import logging
import threading
import time
import webbrowser
from pathlib import Path
import configparser

from webdriver_manager.chrome import ChromeDriverManager

# Variável global para a instância do widget de log (ScrolledText)
log_text_widget = None
# Variável global para a label que exibe o ID do campo ativo
current_id_label = None

APP_NAME = "4keypdv"

DATA_DIR = Path.home() / ".local" / "share" / APP_NAME
DATA_DIR.mkdir(parents=True, exist_ok=True)

LOG_DIR = DATA_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

LOG_FILE = LOG_DIR / "4keypdv.log"

CONFIG_FILE = DATA_DIR / "config.ini"

config = configparser.ConfigParser()

if not CONFIG_FILE.exists():
    config["PDV"] = {
        "PDV_URL": "http://servidor/maxipos_backoffice/app",
        "MAESTRO_URL": "https://login.servidor.com.br",
        "FILIAL": "42",
        "HEADLESS": "True"
    }
    with open(CONFIG_FILE, "w") as f:
        config.write(f)

config.read(CONFIG_FILE)

PDV_URL = config["PDV"]["PDV_URL"]
MAESTRO_URL = config["PDV"]["MAESTRO_URL"]
FILIAL = config["PDV"]["FILIAL"]
HEADLESS = config["PDV"].getboolean("HEADLESS")


class TextHandler(logging.Handler):
    """
    Um manipulador de log personalizado que escreve registros de log em um widget Text do Tkinter.
    """
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
        self.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

    def emit(self, record):
        msg = self.format(record)
        try:
            self.text_widget.after(0, self.insert_message, msg)
        except RuntimeError as e:
            print(f"Logging to console as Tkinter mainloop not active yet: {msg} (Error: {e})")

    def insert_message(self, msg):
        self.text_widget.configure(state='normal')
        self.text_widget.insert(tk.END, msg + '\n')
        self.text_widget.see(tk.END)
        self.text_widget.configure(state='disabled')


def setup_logging(text_widget):
    """Configura o logging para enviar a saída para o console, arquivo e widget Text do Tkinter."""
    global log_text_widget
    log_text_widget = text_widget

    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logging.root.addHandler(console_handler)

    file_handler = logging.FileHandler(LOG_FILE)
    file_handler.setFormatter(
        logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    )
    logging.root.addHandler(file_handler)

    text_handler = TextHandler(text_widget)
    logging.root.addHandler(text_handler)

    logging.root.setLevel(logging.INFO)


def create_driver(headless=True):
    chrome_options = ChromeOptions()
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-popup-blocking")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--window-size=1280,860")

    if headless:
        chrome_options.add_argument("--headless=new")

    service = Service(ChromeDriverManager().install())

    return webdriver.Chrome(service=service, options=chrome_options)


def exibir_mensagem_sucesso(usuario, senha, senha_pdv, nome_usuario_extraido, numero_usuario_extraido, parent_window):
    """
    Exibe uma mensagem de sucesso da automação de alteração de senha e, em seguida,
    salva os dados no arquivo Excel.
    """
    messagebox.showinfo("Sucesso da Automação", f"Senha mudada com sucesso para o usuário: {nome_usuario_extraido} (ID: {numero_usuario_extraido})!", parent=parent_window)
    salvar_dados_excel(usuario, senha, senha_pdv, nome_usuario_extraido, numero_usuario_extraido)


def exibir_mensagem_falha(erro, parent_window):
    """Exibe uma mensagem de falha da automação."""
    messagebox.showerror("Falha", f"Falha na automação. Erro: {erro}", parent=parent_window)


def salvar_dados_excel(usuario, senha, senha_pdv, nome_usuario_extraido, numero_usuario_extraido):
    """
    Salva os dados do usuário, senha, senha PDV e dados extraídos em um arquivo Excel.
    """
    nome_arquivo = DATA_DIR / "4keypdv_history.xlsx"

    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
        sheet = workbook.active
        logging.info(f"Arquivo '{nome_arquivo}' existente. Adicionando dados.")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Data e Hora", "Usuário Login", "Senha Login", "Senha PDV", "Nome do Usuário", "Número do Usuário"])
        logging.info(f"Criado novo arquivo '{nome_arquivo}'.")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([now, usuario, senha, senha_pdv, nome_usuario_extraido, numero_usuario_extraido])

    try:
        workbook.save(nome_arquivo)
        logging.info(f"Dados da senha e informações do usuário salvos com sucesso em '{nome_arquivo}'.")
    except Exception as e:
        logging.error(f"Erro ao salvar os dados no Excel: {e}")
        if log_text_widget:
            log_text_widget.after(0, lambda: messagebox.showerror("Erro ao Salvar Excel", f"Erro ao salvar os dados no Excel: {e}", parent=None))


def criar_novo_excel_confirmado(parent_window):
    """
    Cria um novo arquivo Excel, sobrescrevendo o antigo, após a confirmação do usuário.
    """
    nome_arquivo = DATA_DIR / "4keypdv_history.xlsx"

    if messagebox.askyesno(
        "Confirmar Criação de Novo Excel",
        "Tem certeza que deseja criar um NOVO arquivo Excel?\n"
        "O arquivo existente (4keypdv_history.xlsx) será APAGADO e um novo será criado.",
        parent=parent_window
    ):
        try:
            if os.path.exists(nome_arquivo):
                os.remove(nome_arquivo)
                logging.info(f"Arquivo antigo '{nome_arquivo}' removido com sucesso.")

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Data e Hora", "Usuário Login", "Senha Login", "Senha PDV", "Nome do Usuário", "Número do Usuário"])
            workbook.save(nome_arquivo)
            logging.info(f"Novo arquivo Excel '{nome_arquivo}' criado com sucesso.")
            messagebox.showinfo("Sucesso", "Novo arquivo Excel criado com sucesso!", parent=parent_window)
        except Exception as e:
            logging.error(f"Erro ao criar novo arquivo Excel: {e}")
            messagebox.showerror("Erro", f"Erro ao criar novo arquivo Excel: {e}", parent=parent_window)
    else:
        logging.info("Criação de novo arquivo Excel cancelada pelo usuário.")


def carregar_dados_excel():
    """
    Carrega as últimas credenciais salvas de '4keypdv_history.xlsx'.
    Retorna (usuario, senha, senha_pdv, nome_usuario_extraido, numero_usuario_extraido)
    ou (None, None, None, None, None) se o arquivo não for encontrado ou estiver vazio.
    """
    nome_arquivo = DATA_DIR / "4keypdv_history.xlsx"

    usuario = None
    senha = None
    senha_pdv = None
    nome_usuario_extraido = None
    numero_usuario_extraido = None

    if not os.path.exists(nome_arquivo):
        logging.info(f"Arquivo '{nome_arquivo}' não encontrado. Iniciando sem dados pré-preenchidos.")
        return None, None, None, None, None

    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
        sheet = workbook.active

        if sheet.max_row > 1:
            last_row = sheet[sheet.max_row]
            usuario = last_row[1].value
            senha = last_row[2].value
            senha_pdv = last_row[3].value
            if sheet.max_column > 4:
                nome_usuario_extraido = last_row[4].value
            if sheet.max_column > 5:
                numero_usuario_extraido = last_row[5].value

            logging.info(f"Dados pré-preenchidos carregados do Excel: Usuário={usuario}, Senha={'*' * len(str(senha)) if senha else 'N/A'}, Nome Extraído={nome_usuario_extraido}, Número Extraído={numero_usuario_extraido}")
        else:
            logging.info(f"Arquivo '{nome_arquivo}' encontrado, mas vazio ou apenas com cabeçalho.")

    except Exception as e:
        logging.error(f"Erro ao carregar dados do Excel para pré-preenchimento: {e}")

    return usuario, senha, senha_pdv, nome_usuario_extraido, numero_usuario_extraido


# --- Funções para Credenciais da Carga PDV ---
def salvar_carga_pdv_credentials(usuario, senha):
    """
    Salva as credenciais de login da Carga PDV em um arquivo Excel separado.
    Mantém apenas a última credencial usada.
    """
    nome_arquivo = DATA_DIR / "4keypdv_credentials.xlsx"

    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
        sheet = workbook.active
        logging.info(f"Arquivo de credenciais de Carga PDV '{nome_arquivo}' existente. Atualizando dados.")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Usuário Carga PDV", "Senha Carga PDV"])
        logging.info(f"Criado novo arquivo de credenciais de Carga PDV '{nome_arquivo}'.")

    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)

    sheet.append([usuario, senha])

    try:
        workbook.save(nome_arquivo)
        logging.info(f"Credenciais de Carga PDV salvas com sucesso em '{nome_arquivo}'.")
    except Exception as e:
        logging.error(f"Erro ao salvar credenciais de Carga PDV no Excel: {e}")
        if log_text_widget:
            log_text_widget.after(0, lambda: messagebox.showerror("Erro ao Salvar Credenciais", f"Erro ao salvar credenciais de Carga PDV: {e}", parent=None))


def carregar_carga_pdv_credentials():
    """
    Carrega as últimas credenciais de login da Carga PDV de um arquivo Excel.
    Retorna (usuario, senha) ou (None, None) se o arquivo não for encontrado ou estiver vazio.
    """
    nome_arquivo = DATA_DIR / "4keypdv_credentials.xlsx"

    usuario = None
    senha = None

    if not os.path.exists(nome_arquivo):
        logging.info(f"Arquivo de credenciais de Carga PDV '{nome_arquivo}' não encontrado. Iniciando sem dados pré-preenchidos.")
        return None, None

    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
        sheet = workbook.active

        if sheet.max_row > 1:
            last_row = sheet[sheet.max_row]
            usuario = last_row[0].value
            senha = last_row[1].value
            logging.info(f"Credenciais de Carga PDV pré-preenchidas carregadas: Usuário={usuario}, Senha={'*' * len(str(senha)) if senha else 'N/A'}")
        else:
            logging.info(f"Arquivo de credenciais de Carga PDV '{nome_arquivo}' encontrado, mas vazio ou apenas com cabeçalho.")

    except Exception as e:
        logging.error(f"Erro ao carregar credenciais de Carga PDV para pré-preenchimento: {e}")

    return usuario, senha


# --- Diálogo para Credenciais da Carga PDV ---
def show_carga_pdv_dialog(parent_window, log_display):
    """
    Exibe um diálogo Tkinter para o usuário inserir as credenciais de login da Carga PDV.
    """
    dialog = tk.Toplevel(parent_window)
    dialog.title("Credenciais de Carga PDV")
    dialog.transient(parent_window)
    dialog.grab_set()

    last_usuario, last_senha = carregar_carga_pdv_credentials()

    tk.Label(dialog, text="Usuário Carga PDV:", font=("Arial", 12)).grid(row=0, column=0, padx=10, pady=5, sticky='w')
    usuario_entry = tk.Entry(dialog, width=30, font=("Arial", 12))
    usuario_entry.grid(row=0, column=1, padx=10, pady=5, sticky='ew')
    if last_usuario:
        usuario_entry.insert(0, last_usuario)

    tk.Label(dialog, text="Senha Carga PDV:", font=("Arial", 12)).grid(row=1, column=0, padx=10, pady=5, sticky='w')
    senha_entry = tk.Entry(dialog, width=30, font=("Arial", 12), show="*")
    senha_entry.grid(row=1, column=1, padx=10, pady=5, sticky='ew')
    if last_senha:
        senha_entry.insert(0, last_senha)

    filial_value = FILIAL

    def on_ok():
        usuario_carga = usuario_entry.get()
        senha_carga = senha_entry.get()

        if usuario_carga and senha_carga:
            salvar_carga_pdv_credentials(usuario_carga, senha_carga)

            log_display.configure(state='normal')
            log_display.delete(1.0, tk.END)
            log_display.configure(state='disabled')

            threading.Thread(target=automacao_carga_pdv, args=(usuario_carga, senha_carga, filial_value, parent_window)).start()
            dialog.destroy()
        else:
            messagebox.showwarning("Dados Incompletos", "Por favor, preencha usuário e senha para Carga PDV.", parent=dialog)

    tk.Button(dialog, text="OK", command=on_ok, font=("Arial", 12), padx=10, pady=5).grid(row=2, column=0, columnspan=2, pady=10)

    dialog.update_idletasks()
    x = parent_window.winfo_x() + (parent_window.winfo_width() // 2) - (dialog.winfo_width() // 2)
    y = parent_window.winfo_y() + (parent_window.winfo_height() // 2) - (dialog.winfo_height() // 2)
    dialog.geometry(f"+{x}+{y}")

    parent_window.wait_window(dialog)


# --- Automação de Carga PDV ---
def automacao_carga_pdv(usuario_carga, senha_carga, filial, parent_window):
    """
    Executa a automação Selenium para enviar carga no PDV.
    """
    driver = None
    try:
        logging.info("Iniciando WebDriver para Carga PDV em modo headless.")
        driver = create_driver(HEADLESS)
        logging.info("WebDriver da Carga PDV iniciado com sucesso.")

        driver.get(PDV_URL)
        logging.info(f"Navegando para: {driver.current_url}")

        logging.info("Preenchendo campo de usuário para Carga PDV...")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "username"))).send_keys(usuario_carga)
        logging.info("Preenchendo campo de senha para Carga PDV...")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password"))).send_keys(senha_carga)
        logging.info("Clicando no botão de login para Carga PDV...")
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "kc-login"))).click()
        logging.info("Login de Carga PDV tentado.")

        logging.info(f"Usuário Carga PDV informado: {usuario_carga}")
        logging.info(f"Senha Carga PDV informada: {'*' * len(senha_carga)}")

        logging.info("Aguardando carregamento da página após login de Carga PDV...")
        WebDriverWait(driver, 15).until(EC.url_contains("/maxipos_backoffice/app"))
        logging.info("Página após login de Carga PDV carregada.")

        max_retries = 3
        for attempt in range(max_retries):
            try:
                logging.info(f"Tentativa {attempt + 1}: Clicando para selecionar a filial (primeiro clique)...")
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[12]/div[2]/div[4]"))).click()
                logging.info(f"Tentativa {attempt + 1}: Clicando para selecionar a filial (segundo clique)...")
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[12]/div[2]/span[4]/div[1]"))).click()

                logging.info(f"Tentativa {attempt + 1}: Selecionando opção no dropdown de filial...")
                select_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[2]/select")))
                select_obj = Select(select_element)
                select_obj.select_by_index(2)

                logging.info(f"Tentativa {attempt + 1}: Inserindo número da filial: {filial}")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div[2]/div[2]/input"))).send_keys(str(filial))

                logging.info(f"Tentativa {attempt + 1}: Clicando no botão de confirmar filial...")
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[13]/form/div[1]/div[3]/div[2]/div[1]/div"))).click()

                logging.info(f"Tentativa {attempt + 1}: Aguardando alerta de conclusão da carga...")
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[6]/center/div[1]"))).click()
                logging.info("Carga PDV concluída com sucesso.")
                parent_window.after(0, lambda: messagebox.showinfo("Sucesso da Carga PDV", "Carga PDV enviada com sucesso!", parent=parent_window))
                break
            except StaleElementReferenceException as stale_e:
                logging.warning(f"StaleElementReferenceException na Carga PDV (Tentativa {attempt + 1}/{max_retries}): {stale_e}")
                if attempt == max_retries - 1:
                    raise
                time.sleep(2)

    except Exception as e:
        logging.error(f"Erro na automação de Carga PDV: {e}", exc_info=True)
        parent_window.after(0, lambda error_msg=e: exibir_mensagem_falha(f"Erro na Carga PDV: {error_msg}. Consulte o log para mais detalhes.", parent_window))
    finally:
        if driver:
            driver.quit()
            logging.info("WebDriver da Carga PDV fechado.")


# --- Automação de Alteração de Senha ---
def automacao_selenium(usuario, senha, senha_pdv, parent_window):
    """
    Executa a automação Selenium para alterar a senha do PDV e extrai o nome e número do usuário.
    """
    driver = None
    nome_usuario_extraido = "N/A"
    numero_usuario_extraido = "N/A"

    try:
        logging.info("Iniciando o WebDriver em modo headless.")
        driver = create_driver(HEADLESS)
        logging.info("WebDriver iniciado com sucesso.")

        driver.get(PDV_URL)
        logging.info(f"Navegando para: {driver.current_url}")

        logging.info("Preenchendo campo de usuário...")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "username"))).send_keys(usuario)
        logging.info("Preenchendo campo de senha...")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password"))).send_keys(senha)
        logging.info("Clicando no botão de login...")
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "kc-login"))).click()
        logging.info("Login tentado.")

        logging.info(f"Usuário informado: {usuario}")
        logging.info(f"Senha informada: {'*' * len(senha)}")

        logging.info("Aguardando carregamento da página após login...")
        WebDriverWait(driver, 15).until(EC.url_contains("/maxipos_backoffice/app"))
        logging.info("Página após login carregada.")

        logging.info("Clicando no menu 'Usuário'...")
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='mktmenu_item_nivel1_text' and text()='Usuário']"))).click()
        logging.info("Clicando no submenu 'Alterar Senha'...")
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='mktmenu_item_nivel2_text' and text()='Alterar Senha']"))).click()
        logging.info("Página de alteração de senha carregada.")

        logging.info("Tentando extrair nome e número do usuário...")
        try:
            nome_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[4]/input")))
            nome_usuario_extraido = nome_element.get_attribute("value")
            logging.info(f"Nome do Usuário Extraído: {nome_usuario_extraido}")
        except (NoSuchElementException, TimeoutException):
            logging.warning("Não foi possível extrair o Nome do Usuário. Campo não encontrado ou tempo esgotado.")

        try:
            numero_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[3]/input")))
            numero_usuario_extraido = numero_element.get_attribute("value")
            logging.info(f"Número do Usuário Extraído: {numero_usuario_extraido}")
        except (NoSuchElementException, TimeoutException):
            logging.warning("Não foi possível extrair o Número do Usuário. Campo não encontrado ou tempo esgotado.")

        logging.info("Preenchendo o primeiro campo de senha PDV...")
        campo1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[11]/input")))
        campo1.send_keys(senha_pdv)
        logging.info("Preenchendo o segundo campo de senha PDV...")
        campo2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[14]/input")))
        campo2.send_keys(senha_pdv)
        logging.info("Clicando no botão de salvar...")
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[13]/form/div[1]/div[3]/div[2]/div[1]"))).click()
        logging.info("Botão de salvar clicado. Aguardando confirmação.")

        try:
            WebDriverWait(driver, 5).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            logging.info(f"Alerta presente: {alert.text}")
            alert.accept()
            logging.info("Alerta aceito.")
        except TimeoutException:
            logging.warning("Nenhum alerta de sucesso ou falha apareceu dentro do tempo limite.")

        parent_window.after(0, lambda: exibir_mensagem_sucesso(usuario, senha, senha_pdv, nome_usuario_extraido, numero_usuario_extraido, parent_window))
        logging.info("Automação concluída com sucesso.")

    except StaleElementReferenceException:
        logging.warning("Elemento obsoleto encontrado. Tentando reencontrar e continuar...")
        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='mktmenu_item_nivel1_text' and text()='Usuário']"))).click()
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='mktmenu_item_nivel2_text' and text()='Alterar Senha']"))).click()
            try:
                nome_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[4]/input")))
                nome_usuario_extraido = nome_element.get_attribute("value")
            except (NoSuchElementException, TimeoutException):
                logging.warning("Não foi possível reencontrar e extrair o Nome do Usuário após StaleElementReferenceException.")
            try:
                numero_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[3]/input")))
                numero_usuario_extraido = numero_element.get_attribute("value")
            except (NoSuchElementException, TimeoutException):
                logging.warning("Não foi possível reencontrar e extrair o Número do Usuário após StaleElementReferenceException.")

            campo1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[11]/input")))
            campo1.send_keys(senha_pdv)
            campo2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[14]/input")))
            campo2.send_keys(senha_pdv)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[13]/form/div[1]/div[3]/div[2]/div[1]"))).click()
            parent_window.after(0, lambda: exibir_mensagem_sucesso(usuario, senha, senha_pdv, nome_usuario_extraido, numero_usuario_extraido, parent_window))
            logging.info("Automação recuperada e concluída com sucesso.")
        except (NoSuchElementException, TimeoutException) as reencontrar_erro:
            logging.error(f"Erro ao tentar reencontrar elemento após StaleElementReferenceException: {reencontrar_erro}")
            parent_window.after(0, lambda error_msg=reencontrar_erro: exibir_mensagem_falha(f"Erro ao tentar reencontrar elemento: {error_msg}", parent_window))
    except TimeoutException as e:
        logging.error("Tempo limite excedido. O código parou de esperar por um elemento.")
        parent_window.after(0, lambda error_msg=e: exibir_mensagem_falha(f"Tempo limite excedido. Verifique se a página carregou corretamente ou se os seletores estão corretos. Erro: {error_msg}", parent_window))
    except NoSuchElementException as e:
        logging.error(f"Elemento não encontrado: {e}")
        parent_window.after(0, lambda error_msg=e: exibir_mensagem_falha(f"Elemento não encontrado: {error_msg}. Verifique se o XPATH ou ID do elemento está correto.", parent_window))
    except Exception as e:
        logging.error(f"Erro inesperado na automação: {e}", exc_info=True)
        parent_window.after(0, lambda error_msg=e: exibir_mensagem_falha(f"Erro inesperado: {error_msg}. Consulte o log para mais detalhes.", parent_window))
    finally:
        if driver:
            driver.quit()
            logging.info("WebDriver fechado.")


def automacao_selenium_batch(usuario, senha, senha_pdv):
    """
    Versão da automação para execução em lote.
    Não exibe messageboxes de sucesso/erro, apenas registra no log.
    Retorna True em caso de sucesso, False em caso de falha.
    """
    driver = None
    try:
        logging.info(f"[LOTE/SELEÇÃO] Iniciando WebDriver para o usuário: {usuario}")
        driver = create_driver(HEADLESS)
        driver.get(PDV_URL)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "username"))).send_keys(usuario)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password"))).send_keys(senha)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "kc-login"))).click()

        WebDriverWait(driver, 20).until(EC.url_contains("/maxipos_backoffice/app"))
        logging.info(f"[LOTE/SELEÇÃO] Login para {usuario} bem-sucedido.")

        max_retries = 3
        for attempt in range(max_retries):
            try:
                logging.info(f"[LOTE/SELEÇÃO] Tentativa {attempt + 1}: Aguardando e clicando no menu 'Usuário'.")
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='mktmenu_item_nivel1_text' and text()='Usuário']"))).click()

                logging.info(f"[LOTE/SELEÇÃO] Tentativa {attempt + 1}: Clicando no submenu 'Alterar Senha'.")
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='mktmenu_item_nivel2_text' and text()='Alterar Senha']"))).click()

                nome_usuario_extraido = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[4]/input"))).get_attribute("value")
                numero_usuario_extraido = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[3]/input"))).get_attribute("value")

                campo1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[11]/input")))
                campo1.send_keys(senha_pdv)
                campo2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[13]/form/div[2]/div/div[14]/input")))
                campo2.send_keys(senha_pdv)
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[13]/form/div[1]/div[3]/div[2]/div[1]"))).click()

                try:
                    WebDriverWait(driver, 5).until(EC.alert_is_present()).accept()
                except TimeoutException:
                    logging.warning(f"[LOTE/SELEÇÃO] Nenhum alerta de confirmação para {usuario}.")

                logging.info(f"[LOTE/SELEÇÃO] Senha alterada com sucesso para {nome_usuario_extraido} ({numero_usuario_extraido}).")
                salvar_dados_excel(usuario, senha, senha_pdv, nome_usuario_extraido, numero_usuario_extraido)
                return True

            except StaleElementReferenceException:
                logging.warning(f"[LOTE/SELEÇÃO] Stale Element na tentativa {attempt + 1}/{max_retries} para o usuário {usuario}. Tentando novamente...")
                if attempt == max_retries - 1:
                    logging.error(f"[LOTE/SELEÇÃO] Falha final por Stale Element para {usuario} após {max_retries} tentativas.")
                    raise
                time.sleep(2)

    except TimeoutException as e:
        logging.error(f"[LOTE/SELEÇÃO] Falha de Timeout para o usuário {usuario}. Provável falha de login (senha incorreta no Excel?). Erro: {e.msg}", exc_info=False)
        return False
    except Exception as e:
        logging.error(f"[LOTE/SELEÇÃO] Falha inesperada ao alterar senha para o usuário {usuario}. Erro: {e}", exc_info=True)
        return False
    finally:
        if driver:
            driver.quit()

    return False


def run_automation_for_selected(users_to_update, parent_window, process_name="SELEÇÃO"):
    """Executa a automação sequencialmente para uma lista de usuários selecionados."""
    total_users = len(users_to_update)
    logging.info(f"--- INICIANDO PROCESSO DE ALTERAÇÃO PARA {total_users} USUÁRIOS ({process_name}) ---")
    success_count = 0
    fail_count = 0
    for i, user_data in enumerate(users_to_update):
        _, user, pwd, pwd_pdv, name, _ = user_data
        logging.info(f"Processando usuário {i+1}/{total_users}: {name} (Login: {user})")

        success = automacao_selenium_batch(user, pwd, pwd_pdv)
        if success:
            success_count += 1
        else:
            fail_count += 1
        time.sleep(2)

    def update_log_and_show_summary():
        log_text_widget.configure(state='normal')
        log_text_widget.insert(tk.END, f"\n--- PROCESSO DE {process_name} CONCLUÍDO ---\nSucessos: {success_count}\nFalhas: {fail_count}\n")
        log_text_widget.see(tk.END)
        log_text_widget.configure(state='disabled')
        messagebox.showinfo(
            f"Processo de {process_name} Concluído",
            f"Alteração para usuários finalizada.\n\nSucessos: {success_count}\nFalhas: {fail_count}\n\nConsulte o log para detalhes.",
            parent=parent_window
        )

    if log_text_widget:
        log_text_widget.after(0, update_log_and_show_summary)


def show_history_window(parent_window, usuario_entry, senha_entry, senha_pdv_entry):
    """
    Cria e exibe a janela com o histórico de senhas alteradas.
    """
    history_win = tk.Toplevel(parent_window)
    history_win.title("Histórico de Senhas Alteradas")
    history_win.minsize(640, 500)
    history_win.transient(parent_window)
    history_win.grab_set()

    history_win.checkbox_list = []

    search_frame = tk.Frame(history_win, padx=10, pady=10)
    search_frame.pack(fill=tk.X)
    tk.Label(search_frame, text="Pesquisar por Nome:", font=("Arial", 11)).pack(side=tk.LEFT)
    search_entry = tk.Entry(search_frame, font=("Arial", 11))
    search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

    canvas_frame = tk.Frame(history_win)
    canvas_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    canvas = tk.Canvas(canvas_frame)
    scrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas)

    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    def load_and_filter_history():
        nome_arquivo = DATA_DIR / "4keypdv_history.xlsx"

        if not os.path.exists(nome_arquivo):
            logging.warning("Arquivo '4keypdv_history.xlsx' não encontrado.")
            return []

        try:
            workbook = openpyxl.load_workbook(nome_arquivo)
            sheet = workbook.active

            latest_entries = {}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 6:
                    continue

                dt_str, user, pwd, pwd_pdv, name, num_user = row
                if not name or not num_user:
                    continue

                try:
                    current_dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                except (ValueError, TypeError):
                    continue

                user_key = (name, num_user)

                if user_key in latest_entries:
                    existing_dt_str = latest_entries[user_key][0]
                    existing_dt = datetime.strptime(existing_dt_str, "%Y-%m-%d %H:%M:%S")
                    if current_dt > existing_dt:
                        latest_entries[user_key] = row
                else:
                    latest_entries[user_key] = row

            filtered_data = sorted(list(latest_entries.values()), key=lambda x: x[4])
            return filtered_data

        except Exception as e:
            logging.error(f"Erro ao ler e filtrar o histórico do Excel: {e}")
            return []

    def populate_history_list(search_term=""):
        for widget in scrollable_frame.winfo_children():
            widget.destroy()

        history_win.checkbox_list.clear()
        all_data = load_and_filter_history()

        for data_row in all_data:
            dt_str, user, pwd, pwd_pdv, name, num_user = data_row

            if search_term.lower() not in name.lower():
                continue

            row_frame = tk.Frame(scrollable_frame, bd=1, relief=tk.RIDGE)
            row_frame.pack(fill=tk.X, padx=5, pady=2)

            check_var = tk.BooleanVar()
            history_win.checkbox_list.append((check_var, data_row))

            change_date = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
            is_expired = (datetime.now() - change_date).days > 3
            date_color = "red" if is_expired else "black"

            check_btn = tk.Checkbutton(row_frame, variable=check_var)
            date_label = tk.Label(row_frame, text=f"{change_date.strftime('%d/%m/%Y')}", fg=date_color, font=("Arial", 10, "bold"))
            name_label = tk.Label(row_frame, text=f"{name} (ID: {num_user})", font=("Arial", 10))
            fill_btn = tk.Button(row_frame, text="Preencher", font=("Arial", 9), command=lambda u=user, p=pwd, pp=pwd_pdv: fill_main_entries(u, p, pp))

            check_btn.grid(row=0, column=0, sticky='w', padx=(5, 10))
            date_label.grid(row=0, column=1, sticky='w')
            name_label.grid(row=0, column=2, sticky='w', padx=10)
            fill_btn.grid(row=0, column=3, sticky='e', padx=5)

            row_frame.grid_columnconfigure(0, weight=0)
            row_frame.grid_columnconfigure(1, weight=0)
            row_frame.grid_columnconfigure(2, weight=1)
            row_frame.grid_columnconfigure(3, weight=0)

    def fill_main_entries(user, pwd, pwd_pdv):
        usuario_entry.delete(0, tk.END)
        usuario_entry.insert(0, user)
        senha_entry.delete(0, tk.END)
        senha_entry.insert(0, pwd)
        senha_pdv_entry.delete(0, tk.END)
        senha_pdv_entry.insert(0, pwd_pdv)
        history_win.destroy()

    def start_automation_for_selected_thread():
        selected_users = [data for var, data in history_win.checkbox_list if var.get()]

        if not selected_users:
            messagebox.showwarning("Nenhum Usuário Selecionado", "Por favor, selecione pelo menos um usuário na lista para alterar a senha.", parent=history_win)
            return

        if messagebox.askyesno(
            "Confirmar Alteração para Selecionados",
            f"Você selecionou {len(selected_users)} usuário(s) para alterar a senha.\n"
            "O processo será executado em segundo plano.\n\n"
            "Deseja continuar?",
            parent=history_win
        ):
            history_win.destroy()
            threading.Thread(target=run_automation_for_selected, args=(selected_users, parent_window, "SELEÇÃO")).start()

    def delete_selected_from_excel():
        selected_data = [data for var, data in history_win.checkbox_list if var.get()]

        if not selected_data:
            messagebox.showwarning("Nenhum Usuário Selecionado", "Por favor, selecione pelo menos um usuário para eliminar.", parent=history_win)
            return

        confirm = messagebox.askyesno(
            "Confirmar Exclusão PERMANENTE",
            f"ATENÇÃO!\n\nVocê está prestes a ELIMINAR PERMANENTEMENTE os registros de {len(selected_data)} usuário(s) do arquivo `4keypdv_history.xlsx`.\n\n"
            "Esta ação NÃO PODE SER DESFEITA.\n\nDeseja continuar?",
            icon='warning',
            parent=history_win
        )

        if not confirm:
            logging.info("Exclusão de registros do Excel cancelada pelo usuário.")
            return

        try:
            nome_arquivo = DATA_DIR / "4keypdv_history.xlsx"

            keys_to_delete = {(data[4], str(data[5])) for data in selected_data}

            workbook = openpyxl.load_workbook(nome_arquivo)
            sheet = workbook.active

            rows_to_keep = []
            rows_to_keep.append([cell.value for cell in sheet[1]])

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 6:
                    continue
                current_key = (row[4], str(row[5]))
                if current_key not in keys_to_delete:
                    rows_to_keep.append(row)

            sheet.delete_rows(1, sheet.max_row)

            for row_data in rows_to_keep:
                sheet.append(row_data)

            workbook.save(nome_arquivo)
            logging.info(f"{len(keys_to_delete)} registro(s) foram eliminados com sucesso do arquivo Excel.")

            populate_history_list(search_entry.get())
            messagebox.showinfo("Sucesso", "Registros selecionados foram eliminados com sucesso.", parent=history_win)

        except Exception as e:
            logging.error(f"Erro ao eliminar registros do Excel: {e}")
            messagebox.showerror("Erro de Exclusão", f"Ocorreu um erro ao tentar eliminar os registros: {e}", parent=history_win)

    bottom_frame = tk.Frame(history_win, pady=10)
    bottom_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=10)

    select_all_var = tk.BooleanVar()

    def toggle_all_checkboxes():
        is_checked = select_all_var.get()
        for var, _ in history_win.checkbox_list:
            var.set(is_checked)

    tk.Checkbutton(bottom_frame, text="Selecionar Todos", variable=select_all_var, command=toggle_all_checkboxes, font=("Arial", 10)).pack(side=tk.LEFT)

    right_button_frame = tk.Frame(bottom_frame)
    right_button_frame.pack(side=tk.RIGHT)

    selected_btn = tk.Button(right_button_frame, text="Alterar Selecionados", command=start_automation_for_selected_thread, font=("Arial", 11), bg="#C1E1C1")
    selected_btn.pack(side=tk.LEFT, padx=5)

    delete_btn = tk.Button(right_button_frame, text="Eliminar Selecionados", command=delete_selected_from_excel, font=("Arial", 11), bg="#FF8C8C")
    delete_btn.pack(side=tk.LEFT, padx=5)

    search_entry.bind("<KeyRelease>", lambda e: populate_history_list(search_entry.get()))
    populate_history_list()

    parent_window.wait_window(history_win)


def run_maestro_login_automation(url, username, password):
    """
    Abre uma janela VISÍVEL do Selenium, navega para a URL do Maestro,
    preenche o usuário e senha e clica em login.
    A janela permanece aberta até que o usuário a feche.
    """
    driver = None
    try:
        chrome_options = ChromeOptions()
        chrome_options.add_argument("--window-size=1080,720")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-extensions")

        logging.info(f"Iniciando automação de login no Maestro para o usuário: {username}")
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.get(url)

        logging.info("Preenchendo campo de usuário...")
        user_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "username")))
        user_field.send_keys(username)

        logging.info("Preenchendo campo de senha...")
        pass_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password")))
        pass_field.send_keys(password)

        logging.info("Clicando no botão 'Fazer Login'...")
        login_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "kc-login")))
        login_button.click()
        logging.info("Login realizado com sucesso.")

        while True:
            if not driver.window_handles:
                logging.info("Janela do Maestro fechada pelo usuário (sem handles).")
                break
            time.sleep(1)

    except Exception:
        logging.info("Janela de login do Maestro foi fechada pelo usuário.")


def main_app():
    """Cria e gerencia a interface gráfica principal da aplicação."""
    root = tk.Tk()
    root.withdraw()

    app_window = tk.Toplevel(root)
    app_window.title("4keypdv")
    app_window.minsize(width=800, height=600)

    global current_id_label
    id_display_frame = tk.Frame(app_window, padx=20, pady=5)
    id_display_frame.pack(side=tk.TOP, fill=tk.X)
    tk.Label(id_display_frame, text="ID do Campo Ativo:", font=("Arial", 12)).pack(side=tk.LEFT)
    current_id_label = tk.Label(id_display_frame, text="N/A", font=("Arial", 12, "bold"), fg="blue")
    current_id_label.pack(side=tk.LEFT, padx=5)

    input_frame = tk.Frame(app_window, padx=20, pady=20)
    input_frame.pack(side=tk.TOP, fill=tk.X)

    tk.Label(input_frame, text="Usuário:", font=("Arial", 16)).grid(row=0, column=0, sticky='w', pady=5)
    usuario_entry = tk.Entry(input_frame, width=40, font=("Arial", 14))
    usuario_entry.grid(row=0, column=1, sticky='ew', pady=5)

    tk.Label(input_frame, text="Senha:", font=("Arial", 16)).grid(row=1, column=0, sticky='w', pady=5)
    senha_entry = tk.Entry(input_frame, width=40, font=("Arial", 14))
    senha_entry.grid(row=1, column=1, sticky='ew', pady=5)

    tk.Label(input_frame, text="Senha PDV (4 dígitos):", font=("Arial", 16)).grid(row=2, column=0, sticky='w', pady=5)
    senha_pdv_entry = tk.Entry(input_frame, width=40, font=("Arial", 14))
    senha_pdv_entry.grid(row=2, column=1, sticky='ew', pady=5)

    input_frame.grid_columnconfigure(1, weight=1)

    latest_usuario, latest_senha, latest_senha_pdv, _, _ = carregar_dados_excel()

    app_window.latest_usuario = latest_usuario
    app_window.latest_senha = latest_senha
    app_window.latest_senha_pdv = latest_senha_pdv

    if latest_usuario:
        usuario_entry.insert(0, latest_usuario)
    if latest_senha:
        senha_entry.insert(0, latest_senha)
    if latest_senha_pdv:
        senha_pdv_entry.insert(0, latest_senha_pdv)

    entry_ids = {
        usuario_entry: "Usuário (username)",
        senha_entry: "Senha (password)",
        senha_pdv_entry: "Senha PDV (pdv_password)"
    }

    def on_entry_focus_in(event):
        if current_id_label:
            current_id_label.config(text=entry_ids.get(event.widget, "N/A"))

    usuario_entry.bind("<FocusIn>", on_entry_focus_in)
    senha_entry.bind("<FocusIn>", on_entry_focus_in)
    senha_pdv_entry.bind("<FocusIn>", on_entry_focus_in)

    def tab_auto_fill(event):
        if event.widget == usuario_entry:
            if app_window.latest_senha:
                senha_entry.delete(0, tk.END)
                senha_entry.insert(0, app_window.latest_senha)
            if app_window.latest_senha_pdv:
                senha_pdv_entry.delete(0, tk.END)
                senha_pdv_entry.insert(0, app_window.latest_senha_pdv)
            return "break"

    usuario_entry.bind("<Tab>", tab_auto_fill)

    log_frame = tk.LabelFrame(app_window, text="Log da Automação", padx=10, pady=10)
    log_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=20, pady=10)

    log_display = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, state='disabled', font=("Consolas", 10), bg="black", fg="lime green")
    log_display.pack(fill=tk.BOTH, expand=True)

    setup_logging(log_display)

    def start_automation_in_thread_wrapper():
        usuario = usuario_entry.get()
        senha = senha_entry.get()
        senha_pdv = senha_pdv_entry.get()

        if usuario and senha and senha_pdv:
            log_display.configure(state='normal')
            log_display.delete(1.0, tk.END)
            log_display.configure(state='disabled')

            threading.Thread(target=automacao_selenium, args=(usuario, senha, senha_pdv, app_window)).start()
        else:
            messagebox.showwarning("Dados Incompletos", "Por favor, preencha todos os campos de login para Alterar Senha.", parent=app_window)

    def start_maestro_login_thread():
        user = usuario_entry.get()
        pwd = senha_entry.get()

        if not user or not pwd:
            messagebox.showwarning("Dados Incompletos", "Por favor, preencha os campos 'Usuário' e 'Senha' para usar esta função.", parent=app_window)
            return

        url = MAESTRO_URL

        threading.Thread(
            target=run_maestro_login_automation,
            args=(url, user, pwd),
            daemon=True
        ).start()

    button_controls_frame = tk.Frame(app_window, pady=10)
    button_controls_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=20)

    tk.Button(
        button_controls_frame,
        text="Esqueci minha senha",
        command=start_maestro_login_thread,
        font=("Arial", 12),
        padx=10,
        pady=5,
        bg="#ADD8E6"
    ).pack(side=tk.LEFT, padx=5)

    tk.Button(
        button_controls_frame,
        text="Novo Excel",
        command=lambda: criar_novo_excel_confirmado(app_window),
        font=("Arial", 12),
        padx=10,
        pady=5,
        bg="#FFCCCC"
    ).pack(side=tk.LEFT, padx=5)

    tk.Button(
        button_controls_frame,
        text="Enviar Carga PDV",
        command=lambda: show_carga_pdv_dialog(app_window, log_display),
        font=("Arial", 12),
        padx=10,
        pady=5,
        bg="#CCFFFF"
    ).pack(side=tk.LEFT, padx=5)

    tk.Button(
        button_controls_frame,
        text="Ver Histórico",
        command=lambda: show_history_window(app_window, usuario_entry, senha_entry, senha_pdv_entry),
        font=("Arial", 12),
        padx=10,
        pady=5,
        bg="#D3D3D3"
    ).pack(side=tk.LEFT, padx=5)

    tk.Button(
        button_controls_frame,
        text="Iniciar Automação",
        command=start_automation_in_thread_wrapper,
        font=("Arial", 16),
        padx=20,
        pady=10
    ).pack(side=tk.RIGHT, padx=10)

    app_window.protocol("WM_DELETE_WINDOW", root.destroy)
    root.mainloop()


if __name__ == "__main__":
    main_app()